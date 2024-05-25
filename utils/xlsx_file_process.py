import openpyxl
from .fileconvert import convert_xls_to_xlsx
import unicodedata as ucd
import uuid
import json


def merge(merge_lists):
    merge_all_list = []  # 接收最终内容并返回
    # TODO 遍历合并单元格
    for merge_list in merge_lists:
        # TODO 获取单个合并单元格的起始行(row_min)终止行(row_max)和起始列(col_min)终止列(col_max)
        row_min, row_max, col_min, col_max = merge_list.min_row, merge_list.max_row, merge_list.min_col, merge_list.max_col
        # 这里判断如果合并单元格起始、终止的行和列都不相等，说明合并单元格既合并了行又合并了列，两个for循环依次取出行列位置分别存在x,y中
        if (row_min != row_max and col_min != col_max):
            row_col = [(x, y) for x in range(row_min, row_max+1) for y in range(col_min, col_max+1)]
            merge_all_list.append(row_col)  # 取出的值存在列表中
         # 这里判断如果合并单元格起始、终止行相等，起始、终止列不相等，说明合并单元格只合并了列，所以行不动，只循环取出列的值，存在y中，行可以随意取row_min/row_max 
        elif (row_min==row_max and col_min != col_max):
            row_col = [(row_min, y) for y in range(col_min, col_max + 1)]
            merge_all_list.append(row_col)  # 取出的值存在列表中
         # 这里判断如果合并单元格起始、终止行不相等，起始、终止列相等，说明合并单元格只合并了行，所以列不动，只循环取出行的值，存在x中，列可以随意取col_min/col_max
        elif (row_min != row_max and col_min == col_max):
            row_col = [(x, col_min) for x in range(row_min, row_max + 1)]
            merge_all_list.append(row_col)  # 取出的值存在列表中
    return merge_all_list  # 最终返回列表
    # TODO 得到的是个这样的列表值：[[(2, 1), (3, 1)], [(5, 1), (6, 1)], [(5, 2), (6, 2)], [(5, 3), (6, 3)]]，列表中每个列表表示合并单元格的跨度


def merge_values(sheet, mm_list, *rr):
    for ii in range(0, len(mm_list)):
        if rr in mm_list[ii]:
            value1 = sheet.cell(row=mm_list[ii][0][0], column=mm_list[ii][0][1]).value
            if value1 is not None and isinstance(value1, str):
                return ucd.normalize('NFKC', value1)
            return value1
    else:
        value2 = sheet.cell(*rr).value
        if value2 is not None and isinstance(value2, str):
            return ucd.normalize('NFKC', value2)
        return value2


def xlsx_parser(xlsx_path, fileuuid):
    workbook = openpyxl.load_workbook(filename=xlsx_path)
    sheet_dict = dict()
    text_dict = dict()
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        merge_lists = sheet.merged_cells
        merge_all_list = merge(merge_lists)

        dict_val = dict()
        dict_val['merge_info'] = merge_all_list
        dict_val['cell_info'] = dict()
        for x in range(sheet.min_row, sheet.max_row+1):
            for y in range(sheet.min_column, sheet.max_column+1):
                i = (x, y)
                uuid4_ = str(uuid.uuid4())
                value_ = merge_values(sheet, merge_all_list, *i)
                dict_val['cell_info'][uuid4_] = (i, value_)
                text_dict[uuid4_] = value_
        sheet_dict[sheet_name] = dict_val
    with open(fileuuid, 'w', encoding='utf-8') as fw:
        fw.write(json.dumps(sheet_dict, ensure_ascii=False))

    return fileuuid, text_dict


def parser_xlsx_trans_result(trans_result, fileuuid, xlsxname):
    with open(fileuuid, 'r', encoding='utf-8') as fr:
        sheet_dict = json.loads(fr.read())
    wk = openpyxl.Workbook()
    for sheetname in sheet_dict.keys():
        merge_info = sheet_dict[sheetname]['merge_info']
        cell_info = sheet_dict[sheetname]['cell_info']
        ws = wk.create_sheet(sheetname)
        for key in trans_result.keys():
            if key in cell_info.keys():
                st_row, st_col = cell_info[key][0]
                ws.cell(row=st_row, column=st_col).value = trans_result[key]
        for merge_ in merge_info:
            start_row = min(merge_, key=lambda item: item[0])[0]
            end_row = max(merge_, key=lambda item: item[0])[0]
            start_column = min(merge_, key=lambda item: item[1])[1]
            end_column = max(merge_, key=lambda item: item[1])[1]
            ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)

    if 'Sheet' in wk.sheetnames:
        wk.remove(wk['Sheet'])
    if len(wk.sheetnames) > 0:
        wk.save(xlsxname)

    return xlsxname
